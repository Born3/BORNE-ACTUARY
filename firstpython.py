import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string

def automate_excel(file_name):
    excel_file = pd.read_excel(file_name)
    excel_file[['Gender', 'Product line', 'Total']]
    report_table = excel_file.pivot_table(index='Gender',columns='Product line',values='Total',aggfunc='sum').round(0)
    print(report_table)
    month_and_extension = file_name.split('_')[1]
    report_table.to_excel(f'report_{month_and_extension}', sheet_name='Report', startrow=4)
    # loading workbook and selecting sheet
    wb = load_workbook(f'report_{month_and_extension}')
    sheet = wb['Report']
    min_column = wb.active.min_column
    min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    barchart = BarChart()
    data = Reference(sheet, min_col= min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
    categories = Reference(sheet, min_col= min_column, max_col=max_column, min_row=min_row+1, max_row=max_row)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    #location chart
    sheet.add_chart(barchart, "B12")
    barchart.title = 'Sales by Product line'
    barchart.style = 5 #choose the chart style
    sheet = wb['Report']
    # sum in columns B-G
    import string
    alphabet = list(string.ascii_uppercase)
    excel_alphabet = alphabet[0:max_column]
    print(excel_alphabet)

    for i in excel_alphabet:
        if i!='A':
            sheet[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
            sheet[f'{i}{max_row+1}'].style = 'Currency'
    # adding total label
    sheet[f'{excel_alphabet[0]}{max_row+1}'] = 'Total'
    # getting month name
    month_name = month_and_extension.split('.')[0]
    sheet = wb['Report']
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month_name.title()
    sheet['A1'].font = Font('Georgia', bold=True, size=20)
    sheet['A2'].font = Font('Georgia', bold=True, size=10)
    wb.save(f'report_{month_and_extension}')
    return
